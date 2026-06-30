"""
export_migration_report.py
--------------------------
Exports one Excel sheet per title_suffix with full video migration data.

Usage:
    python export_migration_report.py
    python export_migration_report.py --out "D:/reports/migration_report.xlsx"
"""

import sys, os, argparse
from datetime import datetime
sys.path.insert(0, os.path.dirname(__file__))

from app.database.session import SessionLocal
from app.database.models import Video
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SUFFIXES = {
    "Baby & Childcare":  "(New_Baby_Childcare)",
    "Germanic":          "(New_Germanic)",
    "Romance":           "(New_Romance)",
    "Slavic":            "(New_Slavic)",
}

HEADERS = [
    "ID", "Vimeo ID", "Title", "Vimeo URL", "Folder",
    "Mux Asset ID", "Mux Public Playback ID", "Stream URL",
    "Mux Signed Playback ID", "Mux DRM Playback ID",
    "Captions Count", "Caption Languages",
    "Audio Tracks Count", "Audio Languages",
    "Status", "Source", "Created At",
]

COL_WIDTHS = [6, 28, 52, 38, 10, 36, 36, 54, 36, 36, 14, 30, 16, 30, 12, 10, 22]

HEADER_FILL   = PatternFill("solid", start_color="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT     = Font(name="Arial", size=10)
ALT_FILL      = PatternFill("solid", start_color="D6E4F0")
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=False)
THIN          = Side(style="thin", color="B0B0B0")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_COLORS = {
    "ready":      "C6EFCE",   # green
    "processing": "FFEB9C",   # yellow
    "errored":    "FFC7CE",   # red
    "pending":    "EDEDED",   # grey
}


def add_sheet(wb: Workbook, sheet_name: str, suffix_slug: str, db):
    ws = wb.create_sheet(title=sheet_name)

    # Query
    videos = (
        db.query(Video)
        .filter(Video.vimeo_id.like(f"%_{suffix_slug}"))
        .order_by(Video.id)
        .all()
    )

    # Summary row above headers
    ws.append([f"{sheet_name}  |  Suffix: ({suffix_slug})  |  Total: {len(videos)}  |  "
               f"Ready: {sum(1 for v in videos if v.status=='ready')}  |  "
               f"Processing: {sum(1 for v in videos if v.status=='processing')}  |  "
               f"Errored: {sum(1 for v in videos if v.status=='errored')}"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    summary_cell = ws.cell(1, 1)
    summary_cell.font  = Font(name="Arial", bold=True, size=11, color="1F4E79")
    summary_cell.fill  = PatternFill("solid", start_color="D6E4F0")
    summary_cell.alignment = LEFT
    ws.row_dimensions[1].height = 20

    # Header row
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(2, col_idx)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
    ws.row_dimensions[2].height = 18

    # Data rows
    for row_num, v in enumerate(videos, 3):
        row = [
            v.id,
            v.vimeo_id,
            v.display_title or v.vimeo_title,
            v.vimeo_url,
            v.vimeo_folder_path or "",
            v.mux_asset_id or "",
            v.mux_playback_id or "",
            v.mux_stream_url or "",
            v.mux_signed_playback_id or "",
            v.mux_drm_playback_id or "",
            v.captions_count or 0,
            v.captions_languages or "",
            v.audio_tracks_count or 0,
            v.audio_languages or "",
            v.status or "",
            v.source or "vimeo",
            v.created_at.strftime("%Y-%m-%d %H:%M") if v.created_at else "",
        ]
        ws.append(row)

        alt = (row_num % 2 == 0)
        status_color = STATUS_COLORS.get(v.status or "", None)

        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row_num, col_idx)
            cell.font      = DATA_FONT
            cell.border    = BORDER
            cell.alignment = CENTER if col_idx in (1, 11, 13, 15, 16) else LEFT

            # Status column gets status color, others get alt row shading
            if col_idx == 15 and status_color:
                cell.fill = PatternFill("solid", start_color=status_color)
            elif alt:
                cell.fill = ALT_FILL

        ws.row_dimensions[row_num].height = 15

    # Column widths
    for col_idx, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Auto-filter on header row
    ws.auto_filter.ref = f"A2:{get_column_letter(len(HEADERS))}2"

    return len(videos)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="migration_report.xlsx",
                        help="Output file path (default: migration_report.xlsx)")
    args = parser.parse_args()

    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    with SessionLocal() as db:
        for sheet_name, suffix_slug in SUFFIXES.items():
            count = add_sheet(wb, sheet_name, suffix_slug, db)
            print(f"  {sheet_name}: {count} videos exported")

    wb.save(args.out)
    print(f"\nSaved: {args.out}")


if __name__ == "__main__":
    main()
