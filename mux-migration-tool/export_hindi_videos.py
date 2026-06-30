"""
export_hindi_videos.py
----------------------
Exports all Hindi videos from DB to an Excel file.
Usage:
    python export_hindi_videos.py
Output:
    D:\Icare\icare-video-platform\mux-migration-tool\hindi_videos.xlsx
"""

import os, sys
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
load_dotenv()

from app.database.session import SessionLocal
from app.database.models import Video

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "hindi_videos.xlsx"

# ── Fetch ─────────────────────────────────────────────────────────────────────
with SessionLocal() as db:
    videos = (
        db.query(Video)
        .filter(Video.vimeo_title.ilike("%Hindi%"))
        .order_by(Video.id)
        .all()
    )

print(f"Found {len(videos)} Hindi videos")

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Hindi Videos"

HEADERS = [
    "ID", "Title", "Status", "Mux Asset ID", "Mux Playback ID",
    "Vimeo ID", "Captions Languages", "Captions Count",
    "Audio Languages", "Audio Tracks Count", "Created At"
]

# ── Header style ──────────────────────────────────────────────────────────────
header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=11)
header_fill    = PatternFill("solid", start_color="1F4E79")
header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin           = Side(style="thin", color="BFBFBF")
cell_border    = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font    = header_font
    cell.fill    = header_fill
    cell.alignment = header_align
    cell.border  = cell_border

ws.row_dimensions[1].height = 30

# ── Row styles ────────────────────────────────────────────────────────────────
fill_even  = PatternFill("solid", start_color="EBF3FB")
fill_odd   = PatternFill("solid", start_color="FFFFFF")
body_font  = Font(name="Arial", size=10)
body_align = Alignment(vertical="center")

STATUS_COLORS = {
    "ready":      "C6EFCE",  # green
    "processing": "FFEB9C",  # yellow
    "errored":    "FFC7CE",  # red
}

for row_idx, v in enumerate(videos, 2):
    row_data = [
        v.id,
        v.vimeo_title or "",
        v.status or "",
        v.mux_asset_id or "",
        v.mux_playback_id or "",
        v.vimeo_id or "",
        v.captions_languages or "",
        v.captions_count or 0,
        v.audio_languages or "",
        v.audio_tracks_count or 0,
        str(v.created_at)[:19] if v.created_at else "",
    ]
    fill = fill_even if row_idx % 2 == 0 else fill_odd

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font      = body_font
        cell.alignment = body_align
        cell.border    = cell_border
        # Status column (col 3) — colour by status
        if col_idx == 3:
            status_color = STATUS_COLORS.get(str(value).lower())
            if status_color:
                cell.fill = PatternFill("solid", start_color=status_color)
            else:
                cell.fill = fill
        else:
            cell.fill = fill

# ── Column widths ─────────────────────────────────────────────────────────────
col_widths = {
    1:  8,   # ID
    2:  45,  # Title
    3:  12,  # Status
    4:  42,  # Mux Asset ID
    5:  36,  # Mux Playback ID
    6:  28,  # Vimeo ID
    7:  22,  # Captions Languages
    8:  16,  # Captions Count
    9:  22,  # Audio Languages
    10: 18,  # Audio Tracks Count
    11: 20,  # Created At
}
for col, width in col_widths.items():
    ws.column_dimensions[get_column_letter(col)].width = width

# ── Freeze header + auto-filter ───────────────────────────────────────────────
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ── Summary row ───────────────────────────────────────────────────────────────
last_row = len(videos) + 2
ws.cell(row=last_row, column=1, value="TOTAL").font = Font(name="Arial", bold=True, size=10)
ws.cell(row=last_row, column=1).fill = PatternFill("solid", start_color="D9E1F2")
ws.cell(row=last_row, column=2, value=f'=COUNTA(B2:B{last_row-1})').font = Font(name="Arial", bold=True, size=10)
ws.cell(row=last_row, column=2).fill = PatternFill("solid", start_color="D9E1F2")

# Status breakdown in a separate mini-table (columns M+)
ws["M1"] = "Status Summary"
ws["M1"].font = Font(name="Arial", bold=True, size=11)
ws["M1"].fill = PatternFill("solid", start_color="1F4E79")
ws["M1"].font = Font(name="Arial", bold=True, color="FFFFFF")
ws["N1"] = "Count"
ws["N1"].font = Font(name="Arial", bold=True, color="FFFFFF")
ws["N1"].fill = PatternFill("solid", start_color="1F4E79")
ws.column_dimensions["M"].width = 18
ws.column_dimensions["N"].width = 10

statuses = {}
for v in videos:
    s = (v.status or "unknown").lower()
    statuses[s] = statuses.get(s, 0) + 1

for i, (status, count) in enumerate(sorted(statuses.items()), 2):
    ws.cell(row=i, column=13, value=status)
    ws.cell(row=i, column=14, value=count)
    color = STATUS_COLORS.get(status, "F2F2F2")
    ws.cell(row=i, column=13).fill = PatternFill("solid", start_color=color)
    ws.cell(row=i, column=14).fill = PatternFill("solid", start_color=color)
    ws.cell(row=i, column=13).font = Font(name="Arial", size=10)
    ws.cell(row=i, column=14).font = Font(name="Arial", size=10)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
print(f"\nStatus breakdown:")
for status, count in sorted(statuses.items()):
    print(f"  {status}: {count}")
