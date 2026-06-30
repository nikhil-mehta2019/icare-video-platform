"""
format_eldercare_schema.py
Reformats Eldercare CSV into base44 schema column order.

Column mapping:
  Title                    -> title
  Mux Asset ID             -> muxAssetId
  Mux Signed Playback ID   -> muxPlaybackId  (signed = primary for playback)
  Mux Playback ID (Public) -> dropped
  Mux DRM Playback ID      -> dropped
  id, is_sample            -> dropped
  missing schema fields    -> added as empty

Usage:
    python format_eldercare_schema.py
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_IN   = Path(r"D:\new uploads\Eldercare - American English - Upload - 13062026.csv")
CSV_OUT  = Path(r"D:\new uploads\Eldercare_schema_import.csv")
XLSX_OUT = Path(r"D:\new uploads\Eldercare_schema_import.xlsx")

SCHEMA_COLS = [
    "chapterCode", "courseCode", "moduleCode", "title", "contentType",
    "vimeoUrl", "muxPlaybackId", "muxPlaybackPolicy", "muxPlaybackTokenRequired",
    "muxPlaybackStatus", "muxAssetId", "slidesUrl", "ispringUrl", "ispringUrls",
    "textContent", "resourceUrl", "estimatedMinutes", "isFreePreview",
    "sortOrder", "status",
]

with open(CSV_IN, newline="", encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))

print(f"Loaded {len(rows)} rows")

# Remap each row to schema
def remap(r):
    signed = r.get("Mux Signed Playback ID", "").strip()
    public = r.get("Mux Playback ID (Public)", "").strip()
    policy = r.get("muxPlaybackPolicy", "public").strip().lower()
    # Use signed if available, else public
    playback_id = signed or public

    return {
        "chapterCode":             r.get("chapterCode", ""),
        "courseCode":              r.get("courseCode", ""),
        "moduleCode":              r.get("moduleCode", ""),
        "title":                   r.get("Title", ""),
        "contentType":             r.get("contentType", ""),
        "vimeoUrl":                "",
        "muxPlaybackId":           playback_id,
        "muxPlaybackPolicy":       r.get("muxPlaybackPolicy", "public"),
        "muxPlaybackTokenRequired": r.get("muxPlaybackTokenRequired", "false"),
        "muxPlaybackStatus":       "ready" if playback_id else "unknown",
        "muxAssetId":              r.get("Mux Asset ID", ""),
        "slidesUrl":               "",
        "ispringUrl":              "",
        "ispringUrls":             r.get("ispringUrls", ""),
        "textContent":             "",
        "resourceUrl":             "",
        "estimatedMinutes":        "",
        "isFreePreview":           "false",
        "sortOrder":               r.get("sortOrder", ""),
        "status":                  r.get("status", "DRAFT").upper(),
    }

remapped = [remap(r) for r in rows]

# CSV
with open(CSV_OUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=SCHEMA_COLS)
    writer.writeheader()
    writer.writerows(remapped)
print(f"CSV saved : {CSV_OUT}")

# Excel
wb = Workbook()
ws = wb.active
ws.title = "Eldercare Import"

thin     = Side(style="thin", color="D0D0D0")
border   = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
hdr_fill = PatternFill("solid", start_color="1F4E79")
body_fnt = Font(name="Arial", size=10)
left_aln = Alignment(horizontal="left", vertical="center")
ctr_aln  = Alignment(horizontal="center", vertical="center")
f_even   = PatternFill("solid", start_color="F2F2F2")
f_odd    = PatternFill("solid", start_color="FFFFFF")

for ci, col in enumerate(SCHEMA_COLS, 1):
    cell           = ws.cell(row=1, column=ci, value=col)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = ctr_aln
    cell.border    = border
ws.row_dimensions[1].height = 28
ws.freeze_panes = "A2"

for ri, r in enumerate(remapped, 2):
    rfill = f_even if ri % 2 == 0 else f_odd
    for ci, col in enumerate(SCHEMA_COLS, 1):
        cell           = ws.cell(row=ri, column=ci, value=r.get(col, ""))
        cell.font      = body_fnt
        cell.alignment = left_aln
        cell.border    = border
        cell.fill      = rfill
    ws.row_dimensions[ri].height = 18

COL_WIDTHS = {
    "chapterCode": 14, "courseCode": 22, "moduleCode": 14, "title": 50,
    "contentType": 14, "vimeoUrl": 20, "muxPlaybackId": 46,
    "muxPlaybackPolicy": 16, "muxPlaybackTokenRequired": 22,
    "muxPlaybackStatus": 16, "muxAssetId": 46, "sortOrder": 10,
    "status": 12, "estimatedMinutes": 16, "isFreePreview": 12, "ispringUrls": 30,
}
for ci, col in enumerate(SCHEMA_COLS, 1):
    ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 14)

ws.auto_filter.ref = f"A1:{get_column_letter(len(SCHEMA_COLS))}1"
wb.save(XLSX_OUT)
print(f"Excel saved: {XLSX_OUT}")
