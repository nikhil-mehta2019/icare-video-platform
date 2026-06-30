"""
create_chapter_mapping.py
-------------------------
Matches chapters from Chapter_export.csv to DB videos migrated in June 2026.
Joins on vimeo_id extracted from vimeoUrl.

Output: chapter_mapping.xlsx

Usage:
    python create_chapter_mapping.py
"""

import os, sys, re, csv
from pathlib import Path
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from dotenv import load_dotenv
load_dotenv()

from app.database.session import SessionLocal
from app.database.models import Video

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CSV_PATH = Path(r"D:\new uploads\Chapter_export.csv")
OUTPUT   = Path(__file__).parent / "chapter_mapping.xlsx"
CUTOFF   = datetime(2026, 6, 1)


def extract_vimeo_id(url: str) -> str | None:
    m = re.search(r'vimeo\.com/(\d+)', url or "")
    return m.group(1) if m else None

def norm_title(t: str) -> str:
    """Lowercase, strip punctuation/spaces for fuzzy title matching."""
    return re.sub(r'[^a-z0-9]', '', (t or "").lower())


# ── Load CSV ──────────────────────────────────────────────────────────────────
with open(CSV_PATH, newline="", encoding="utf-8") as f:
    chapters = list(csv.DictReader(f))

print(f"CSV: {len(chapters)} chapters")

# ── Load DB videos (June 2026) ────────────────────────────────────────────────
with SessionLocal() as db:
    videos = db.query(Video).filter(Video.created_at > CUTOFF).all()

print(f"DB : {len(videos)} videos since {CUTOFF.date()}")

# Build lookups: vimeo_id → Video  AND  normalized_title → Video
vimeo_id_map:  dict[str, Video] = {}
title_map:     dict[str, Video] = {}

for v in videos:
    vid = extract_vimeo_id(v.vimeo_url or "")
    if vid:
        vimeo_id_map[vid] = v
    if v.vimeo_id and v.vimeo_id.isdigit():
        vimeo_id_map[v.vimeo_id] = v
    # title index — both raw and normalized
    raw = (v.vimeo_title or "").strip().lower()
    title_map[raw] = v
    title_map[norm_title(v.vimeo_title or "")] = v

print(f"DB : {len(vimeo_id_map)} videos indexed by vimeo_id")
print(f"DB : {len(title_map)} title index entries\n")

# ── Build rows ────────────────────────────────────────────────────────────────
MATCHED        = "✅ Matched"
TITLE_MATCHED  = "🔍 Title Match"
UPDATED        = "🔄 Needs Update"
NO_MUX         = "⚠️ No Mux ID"
UNMATCHED      = "❌ Not in DB"
ISPRING        = "— iSpring"

rows = []
for ch in chapters:
    content_type = ch.get("contentType", "")
    if content_type != "MUX_VIDEO":
        rows.append({**ch, "_db": None, "_status": ISPRING, "_vimeo_id": "", "_match_method": ""})
        continue

    vimeo_url = ch.get("vimeoUrl", "")
    vid       = extract_vimeo_id(vimeo_url)
    db_video  = vimeo_id_map.get(vid) if vid else None
    match_method = "vimeo_id" if db_video else ""

    # Fallback: match by chapter title against DB vimeo_title
    if not db_video:
        ch_title  = ch.get("title", "")
        db_video  = (title_map.get(ch_title.strip().lower()) or
                     title_map.get(norm_title(ch_title)))
        if db_video:
            match_method = "title"

    if not db_video:
        status = UNMATCHED
    elif not ch.get("muxPlaybackId", "").strip():
        status = NO_MUX if match_method == "vimeo_id" else TITLE_MATCHED
    elif (ch.get("muxAssetId", "") != (db_video.mux_asset_id or "") or
          ch.get("muxPlaybackId", "") != (db_video.mux_signed_playback_id or db_video.mux_playback_id or "")):
        status = UPDATED
    else:
        status = MATCHED

    rows.append({**ch, "_db": db_video, "_status": status,
                 "_vimeo_id": vid or "", "_match_method": match_method})

# ── Stats ─────────────────────────────────────────────────────────────────────
from collections import Counter
status_counts = Counter(r["_status"] for r in rows)
print("Match summary:")
for s, c in sorted(status_counts.items()):
    print(f"  {s}: {c}")
print()

# ── Build Excel ───────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Chapter Mapping"

HEADERS = [
    # Chapter info
    "Chapter Code", "Module Code", "Course Code", "Title", "Content Type",
    "Sort Order",
    # CSV state
    "Vimeo URL", "Vimeo ID",
    "CSV Mux Asset ID", "CSV Mux Playback ID", "CSV Policy",
    # DB match
    "DB Video ID", "DB Title", "DB Status",
    "DB Mux Asset ID", "DB Mux Playback ID (public)",
    "DB Mux Signed Playback ID", "DB Mux DRM Playback ID", "DB Created At",
    # Suggested values to use in CMS
    "➡ Use Mux Playback ID", "➡ Use Mux DRM ID",
    # Action
    "Match Status", "Match Method",
]

thin      = Side(style="thin", color="D0D0D0")
border    = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
body_font = Font(name="Arial", size=10)
ctr       = Alignment(horizontal="center", vertical="center")
left      = Alignment(horizontal="left",   vertical="center", wrap_text=False)

# Header groups
GROUP_FILLS = {
    "chapter": "1F4E79",   # dark blue — chapter info
    "csv":     "375623",   # dark green — CSV state
    "db":      "7B3B00",   # dark brown — DB
    "status":  "4A235A",   # dark purple — match status
}

def hfill(color): return PatternFill("solid", start_color=color)

group_ranges = {
    "chapter": range(1, 7),
    "csv":     range(7, 12),
    "db":      range(12, 20),
    "suggest": range(20, 22),
    "status":  range(22, 24),
}
GROUP_FILLS["suggest"] = "7030A0"  # purple — suggested values

for group, cols in group_ranges.items():
    for col in cols:
        cell = ws.cell(row=1, column=col, value=HEADERS[col-1])
        cell.font      = hdr_font
        cell.fill      = hfill(GROUP_FILLS[group])
        cell.alignment = ctr
        cell.border    = border

ws.row_dimensions[1].height = 32
ws.freeze_panes = "A2"

# Status → row fill
STATUS_FILL = {
    MATCHED:       "E2EFDA",
    TITLE_MATCHED: "DDEBF7",
    UPDATED:       "FFF2CC",
    NO_MUX:        "FCE4D6",
    UNMATCHED:     "FFDCE1",
    ISPRING:       "F2F2F2",
}

for row_idx, r in enumerate(rows, 2):
    db  = r["_db"]
    sfill = PatternFill("solid", start_color=STATUS_FILL.get(r["_status"], "FFFFFF"))

    # "Suggested" columns: signed playback ID (or public if no signed) + DRM
    policy = (r.get("muxPlaybackPolicy") or "").lower()
    suggest_playback = ""
    suggest_drm      = ""
    if db:
        if policy == "signed":
            suggest_playback = db.mux_signed_playback_id or db.mux_playback_id or ""
        else:
            suggest_playback = db.mux_playback_id or db.mux_signed_playback_id or ""
        suggest_drm = db.mux_drm_playback_id or ""

    row_data = [
        r.get("chapterCode", ""),
        r.get("moduleCode", ""),
        r.get("courseCode", ""),
        r.get("title", ""),
        r.get("contentType", ""),
        r.get("sortOrder", ""),
        r.get("vimeoUrl", ""),
        r["_vimeo_id"],
        r.get("muxAssetId", ""),
        r.get("muxPlaybackId", ""),
        r.get("muxPlaybackPolicy", ""),
        db.id                          if db else "",
        db.vimeo_title                 if db else "",
        db.status                      if db else "",
        db.mux_asset_id                if db else "",
        db.mux_playback_id             if db else "",
        db.mux_signed_playback_id      if db else "",
        db.mux_drm_playback_id         if db else "",
        str(db.created_at)[:10]        if db else "",
        suggest_playback,
        suggest_drm,
        r["_status"],
        r["_match_method"],
    ]

    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border    = border
        cell.fill      = sfill
        # Suggested columns (20, 21) — bold + slightly tinted
        if col_idx in (20, 21) and val:
            cell.font      = Font(name="Arial", size=10, bold=True)
            cell.alignment = left
            if val:
                cell.fill = PatternFill("solid", start_color="EAD1F5")  # light purple
        else:
            cell.font      = body_font
            cell.alignment = left

    ws.row_dimensions[row_idx].height = 18

# ── Column widths ─────────────────────────────────────────────────────────────
widths = [14,14,22,48,14,10, 45,14, 38,38,10, 10,48,12, 38,38,38,38,14, 40,38, 20,14]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Auto-filter ───────────────────────────────────────────────────────────────
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")
ws2["A1"] = "Match Status"
ws2["B1"] = "Count"
ws2["A1"].font = Font(name="Arial", bold=True)
ws2["B1"].font = Font(name="Arial", bold=True)
ws2["A1"].fill = hfill("1F4E79")
ws2["A1"].font = Font(name="Arial", bold=True, color="FFFFFF")
ws2["B1"].fill = hfill("1F4E79")
ws2["B1"].font = Font(name="Arial", bold=True, color="FFFFFF")

for i, (s, c) in enumerate(sorted(status_counts.items()), 2):
    ws2.cell(row=i, column=1, value=s).font = Font(name="Arial", size=10)
    ws2.cell(row=i, column=2, value=c).font  = Font(name="Arial", size=10)
    fill = PatternFill("solid", start_color=STATUS_FILL.get(s, "FFFFFF"))
    ws2.cell(row=i, column=1).fill = fill
    ws2.cell(row=i, column=2).fill = fill

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 10

i_total = len(rows)
ws2.cell(row=i+1, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
ws2.cell(row=i+1, column=2, value=i_total).font = Font(name="Arial", bold=True)

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
