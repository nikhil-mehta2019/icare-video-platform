import asyncio
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database.session import SessionLocal
from app.database.models import MigrationJob, MigrationError
from app.services.migration_service import run_bulk_migration
from app.services.report_service import generate_migration_excel
from app.schemas.response_models import MigrationResponse
from app.schemas.request_models import BulkMigrationRequest
from app.services.mux_service import get_all_assets, delete_asset, add_public_playback_id, add_signed_playback_id, delete_playback_id
from app.config import DRM_CONFIGURATION_ID
from app.services.migration_service import process_single_video
from app.services.audio_service import attach_audio_tracks_background
from typing import Optional
from app.services.migration_service import run_folder_migration, run_ids_migration
from typing import List

router = APIRouter(prefix="/migration", tags=["Migration"])

# In-memory progress tracking for background tasks
_task_status = {}

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.post("/vimeo-account", response_model=MigrationResponse)
async def start_migration(
    background_tasks: BackgroundTasks,
    request: BulkMigrationRequest = BulkMigrationRequest(), # Defaults to no limit
    db: Session = Depends(get_db)
):
    # Migration Lock / Idempotency Check
    existing_job = db.query(MigrationJob).filter(MigrationJob.status == "running").first()
    if existing_job:
        raise HTTPException(
            status_code=400, 
            detail=f"A migration is already in progress (Job ID: {existing_job.id}). Please wait for it to complete."
        )

    job = MigrationJob()
    db.add(job)
    db.commit()
    db.refresh(job)

    # Replaced BackgroundTasks with standard asyncio.create_task for reliable async execution
    #asyncio.create_task(run_bulk_migration(job.id, request.limit))
    background_tasks.add_task(run_bulk_migration, job.id, request.limit)

    return {"status": "Migration started", "job_id": job.id}

@router.get("/export")
def export_migration_report():
    excel_file = generate_migration_excel()
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vimeo_mux_mapping.xlsx"}
    )

@router.get("/errors/{job_id}")
def get_migration_errors(job_id: int, db: Session = Depends(get_db)):
    """Return all failed videos for a migration job."""
    job = db.query(MigrationJob).filter(MigrationJob.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Migration job not found")
    errors = db.query(MigrationError).filter(MigrationError.job_id == job_id).all()
    return {
        "job_id": job_id,
        "failed_count": len(errors),
        "errors": [
            {
                "vimeo_id": e.vimeo_id,
                "error_message": e.error_message,
                "failed_at": e.created_at.strftime("%Y-%m-%d %H:%M:%S") if e.created_at else None,
            }
            for e in errors
        ],
    }

@router.get("/status/{job_id}")
def get_migration_status(job_id: int, db: Session = Depends(get_db)):
    """Check the real-time progress of a specific migration job."""
    job = db.query(MigrationJob).filter(MigrationJob.id == job_id).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="Migration job not found")
        
    return {
        "job_id": job.id,
        "status": job.status,           # "running", "completed", or "failed"
        "total_videos": job.total_videos,
        "imported_videos": job.imported_videos,
        "failed_videos": job.failed_videos,
        "percent_complete": round((job.imported_videos + job.failed_videos) / job.total_videos * 100, 1) if job.total_videos > 0 else 0
    }

@router.post("/{job_id}/cancel")
def cancel_migration(job_id: int, db: Session = Depends(get_db)):
    """Stops an active migration job by updating its status."""
    job = db.query(MigrationJob).filter(MigrationJob.id == job_id).first()
    
    if not job:
        raise HTTPException(status_code=404, detail="Migration job not found")
        
    if job.status != "running":
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot cancel job because it is already {job.status}"
        )
        
    # Send the cancellation signal to the database
    job.status = "cancelled"
    db.commit()
    
    return {"status": "success", "message": f"Migration job {job_id} has been cancelled."}

@router.post("/make-public")
async def make_all_assets_public(db: Session = Depends(get_db)):
    """Adds a public playback ID to all existing signed Mux assets and updates the DB."""
    from app.database.models import Video
    videos = db.query(Video).filter(Video.mux_asset_id != None, Video.mux_playback_id != None).all()

    updated, failed = 0, 0
    results = []

    for video in videos:
        try:
            new_playback_id = await asyncio.to_thread(add_public_playback_id, video.mux_asset_id)
            video.mux_playback_id = new_playback_id
            video.mux_stream_url = f"https://stream.mux.com/{new_playback_id}.m3u8"
            db.commit()
            updated += 1
            results.append({"vimeo_id": video.vimeo_id, "status": "updated", "public_playback_id": new_playback_id})
        except Exception as e:
            failed += 1
            results.append({"vimeo_id": video.vimeo_id, "status": "failed", "error": str(e)})

    return {"updated": updated, "failed": failed, "results": results}

@router.delete("/cleanup-mux")
async def cleanup_all_mux_assets():
    """DANGER: Deletes ALL assets from the connected Mux account one by one."""
    try:
        # 1. Fetch all assets from Mux
        assets = await asyncio.to_thread(get_all_assets)
        total_assets = len(assets)
        
        if total_assets == 0:
            return {"status": "success", "message": "No assets found in Mux to delete."}

        deleted_count = 0
        failed_count = 0

        # 2. Loop through and delete them one by one
        for index, asset in enumerate(assets, start=1):
            asset_id = asset["id"]
            try:
                await asyncio.to_thread(delete_asset, asset_id)
                deleted_count += 1
            except Exception as e:
                failed_count += 1
            
            # Pause for 200ms to respect Mux API rate limits
            await asyncio.sleep(0.2)

        return {
            "status": "success",
            "message": f"Mux cleanup complete.",
            "total_found": total_assets,
            "deleted": deleted_count,
            "failed": failed_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/add-signed-playback")
async def backfill_signed_playback_ids(db: Session = Depends(get_db)):
    """
    One-time backfill: adds a signed playback ID to every video that doesn't have one yet.
    Run this once for videos migrated before the download feature was added.
    """
    from app.database.models import Video
    videos = db.query(Video).filter(
        Video.mux_asset_id != None,
        Video.mux_signed_playback_id == None
    ).all()

    updated, failed = 0, 0
    results = []

    for video in videos:
        try:
            new_id, policy_type = await asyncio.to_thread(add_signed_playback_id, video.mux_asset_id)
            video.mux_drm_playback_id = new_id
            db.commit()
            updated += 1
            results.append({"vimeo_id": video.vimeo_id, "status": "updated", "playback_id": new_id, "policy": policy_type})
        except Exception as e:
            failed += 1
            results.append({"vimeo_id": video.vimeo_id, "status": "failed", "error": str(e)})

    return {"updated": updated, "failed": failed, "results": results}


@router.post("/remigrate/{vimeo_id}")
async def remigrate_single_video(vimeo_id: str, db: Session = Depends(get_db)):
    """
    Re-migrates a single video: deletes the old Mux asset, clears the DB record,
    and re-processes from Vimeo. Use this for videos migrated without audio tracks.
    """
    from app.database.models import Video

    video = db.query(Video).filter(Video.vimeo_id == vimeo_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found in database.")

    # Delete old Mux asset if it exists
    if video.mux_asset_id:
        try:
            await asyncio.to_thread(delete_asset, video.mux_asset_id)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Failed to delete old Mux asset: {str(e)}")

    # Save what we need before clearing
    title = video.vimeo_title
    vimeo_url = video.vimeo_url
    folder_path = video.vimeo_folder_path

    # Remove old DB record so process_single_video can create a fresh one
    db.delete(video)
    db.commit()

    # Re-process: fetches from Vimeo and uploads to Mux with audio tracks enabled
    try:
        result = await asyncio.to_thread(
            process_single_video, db, title, vimeo_url, vimeo_id, folder_path
        )
        return {"status": "success", "vimeo_id": vimeo_id, "mux_result": result}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Re-migration failed: {str(e)}")


@router.post("/attach-audio/{vimeo_id}")
async def attach_audio(vimeo_id: str, background_tasks: BackgroundTasks, db: Session = Depends(get_db), language: Optional[str] = None):
    """
    Manually re-triggers audio attachment for an already-migrated video.
    Use ?language=es to attach only a specific language and avoid duplicates.
    """
    from app.database.models import Video
    video = db.query(Video).filter(Video.vimeo_id == vimeo_id).first()
    if not video:
        raise HTTPException(status_code=404, detail="Video not found in database.")
    if not video.mux_asset_id:
        raise HTTPException(status_code=400, detail="Video has no Mux asset ID.")
    if not video.vimeo_url:
        raise HTTPException(status_code=400, detail="Video has no Vimeo URL stored.")

    # Strip suffix from vimeo_id for yt-dlp filename (suffix is DB-only, not a real Vimeo ID)
    raw_vimeo_id = video.vimeo_id.split("_")[0] if "_" in video.vimeo_id else video.vimeo_id

    background_tasks.add_task(
        attach_audio_tracks_background,
        video.mux_asset_id,
        raw_vimeo_id,
        video.vimeo_url,
        language,
    )
    return {"status": "queued", "vimeo_id": vimeo_id, "mux_asset_id": video.mux_asset_id, "language": language or "all"}


async def _run_bulk_audio_attachment(suffix: str, limit: int = None):
    """Background task — attaches audio tracks to all videos whose vimeo_id ends with suffix."""
    import logging
    import os
    from sqlalchemy import text

    BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    LOGS_DIR = os.path.join(BASE_DIR, "logs")
    os.makedirs(LOGS_DIR, exist_ok=True)

    log = logging.getLogger("bulk_audio")
    if not log.handlers:
        formatter = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
        fh = logging.FileHandler(os.path.join(LOGS_DIR, "bulk_audio_attachment.log"), encoding="utf-8")
        fh.setFormatter(formatter)
        fh.setLevel(logging.INFO)
        log.addHandler(fh)
        log.setLevel(logging.INFO)

    log.info(f"{'='*60}")
    log.info(f"[Bulk Audio] Starting — suffix='{suffix}' | limit={limit or 'all'}")

    with SessionLocal() as db:
        rows = db.execute(
            text("SELECT vimeo_id, mux_asset_id, vimeo_url FROM videos WHERE vimeo_id LIKE :pattern AND mux_asset_id IS NOT NULL"),
            {"pattern": f"%{suffix}"}
        ).fetchall()

    if limit:
        rows = rows[:limit]

    total = len(rows)
    log.info(f"[Bulk Audio] Videos to process: {total}")
    _task_status["bulk_audio"] = {
        "status": "running", "total": total,
        "attached": 0, "skipped": 0, "failed": 0, "current": 0, "current_vimeo_id": ""
    }

    attached, skipped, failed = 0, 0, 0

    for i, (vimeo_id, mux_asset_id, vimeo_url) in enumerate(rows, start=1):
        raw_vimeo_id = vimeo_id.split("_")[0] if "_" in vimeo_id else vimeo_id
        _task_status["bulk_audio"].update({"current": i, "current_vimeo_id": vimeo_id})

        log.info(f"[Bulk Audio] --- {i}/{total} | DB vimeo_id: {vimeo_id} | raw: {raw_vimeo_id} | asset: {mux_asset_id}")
        log.info(f"[Bulk Audio]   Vimeo URL: {vimeo_url}")

        try:
            # Capture audio track count before and after by wrapping the call
            # attach_audio_tracks_background logs its own per-track details
            await attach_audio_tracks_background(mux_asset_id, raw_vimeo_id, vimeo_url)

            # Update DB audio_tracks_count based on what actually got attached
            # (The webhook already does this on asset.ready, but we refresh here for bulk runs)
            with SessionLocal() as db:
                db.execute(
                    text("UPDATE videos SET audio_tracks_count = audio_tracks_count WHERE vimeo_id = :vid"),
                    {"vid": vimeo_id}
                )
                db.commit()

            attached += 1
            log.info(f"[Bulk Audio]   ✅ Done for {vimeo_id}")

        except Exception as e:
            failed += 1
            log.error(f"[Bulk Audio]   ❌ Exception for {vimeo_id}: {e}")

        _task_status["bulk_audio"].update({"attached": attached, "skipped": skipped, "failed": failed})
        log.info(f"[Bulk Audio] Progress: {attached} attached | {failed} failed | {total - i} remaining")

    _task_status["bulk_audio"]["status"] = "done"
    log.info(f"[Bulk Audio] {'='*60}")
    log.info(f"[Bulk Audio] FINISHED — Total: {total} | Attached: {attached} | Failed: {failed}")


@router.post("/attach-audio-bulk")
async def attach_audio_bulk(suffix: str = "_052026", limit: Optional[int] = None):
    """
    Triggers audio attachment for videos whose vimeo_id ends with the given suffix.
    Pass ?limit=10 to test on a small batch first.
    Runs sequentially in the background. Poll /migration/task-status for progress.
    Detailed logs written to logs/bulk_audio_attachment.log on the server.
    """
    if _task_status.get("bulk_audio", {}).get("status") == "running":
        raise HTTPException(status_code=400, detail="Bulk audio attachment is already running.")

    asyncio.create_task(_run_bulk_audio_attachment(suffix, limit))
    return {
        "status": "queued",
        "suffix": suffix,
        "limit": limit or "all",
        "message": "Bulk audio attachment started. Poll /migration/task-status or check logs/bulk_audio_attachment.log on server.",
    }


@router.post("/folder-migration")
async def start_folder_migration(
    folder_url: str,
    limit: Optional[int] = None,
    title_suffix: Optional[str] = None,
    background_tasks: BackgroundTasks = BackgroundTasks(),
    db: Session = Depends(get_db)
):
    # Extract folder ID from URL (e.g., .../folder/28548971 -> 28548971)
    try:
        folder_id = folder_url.split("/folder/")[-1].split("?")[0]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Vimeo folder URL")

    # Prevent concurrent migrations
    if db.query(MigrationJob).filter(MigrationJob.status == "running").first():
        raise HTTPException(status_code=400, detail="A migration is already in progress.")

    job = MigrationJob(status="running")
    db.add(job)
    db.commit()
    db.refresh(job)

    background_tasks.add_task(run_folder_migration, job.id, folder_url, limit, title_suffix)

    return {"status": "Folder migration started", "job_id": job.id, "folder_id": folder_id, "title_suffix": title_suffix}


@router.delete("/cleanup-old")
async def cleanup_old_videos(suffix: str, db: Session = Depends(get_db)):
    """
    Deletes all Mux assets and DB records for videos whose title does NOT end with `suffix`.
    Use after re-migrating with a new suffix (e.g. _052026) to remove the old batch.
    Example: DELETE /migration/cleanup-old?suffix=_052026
    """
    from app.database.models import Video

    old_videos = db.query(Video).filter(~Video.vimeo_title.like(f"%{suffix}")).all()

    if not old_videos:
        return {"status": "nothing_to_delete", "deleted": 0, "failed": 0}

    deleted, failed = 0, 0
    results = []

    for video in old_videos:
        if video.mux_asset_id:
            try:
                await asyncio.to_thread(delete_asset, video.mux_asset_id)
            except Exception as e:
                failed += 1
                results.append({"vimeo_id": video.vimeo_id, "status": "failed", "error": str(e)})
                continue

        db.delete(video)
        db.commit()
        deleted += 1
        results.append({"vimeo_id": video.vimeo_id, "title": video.vimeo_title, "status": "deleted"})

    return {"deleted": deleted, "failed": failed, "results": results}


@router.post("/migrate-ids")
async def migrate_ids(
    vimeo_ids: List[str],
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
):
    """
    Migrate a specific list of Vimeo IDs. Already-migrated IDs are skipped automatically.
    Pass Vimeo IDs as a JSON array: ["123456", "789012", ...]
    """
    if not vimeo_ids:
        raise HTTPException(status_code=400, detail="vimeo_ids list is empty.")

    if db.query(MigrationJob).filter(MigrationJob.status == "running").first():
        raise HTTPException(status_code=400, detail="A migration is already in progress.")

    job = MigrationJob(status="running", total_videos=len(vimeo_ids))
    db.add(job)
    db.commit()
    db.refresh(job)

    background_tasks.add_task(run_ids_migration, job.id, vimeo_ids)

    return {"status": "Migration started", "job_id": job.id, "requested": len(vimeo_ids)}


@router.get("/verify-folder")
async def verify_folder_migration(folder_url: str, db: Session = Depends(get_db)):
    """
    Fetches all videos from a Vimeo folder (including sub-folders), checks migration
    status against the DB, and returns an Excel file with one sheet per Vimeo folder name.
    """
    from app.database.models import Video
    from app.services.vimeo_service import get_vimeo_folder_videos
    from io import BytesIO
    import pandas as pd

    try:
        folder_id = folder_url.split("/folder/")[-1].split("?")[0]
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid Vimeo folder URL")

    try:
        all_videos = await asyncio.to_thread(get_vimeo_folder_videos, folder_id)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Vimeo folder: {str(e)}")

    rows = []
    for item in all_videos:
        v = item["video"]
        vimeo_id = v["uri"].split("/")[-1]
        folder_name = item.get("folder_name", "Root")
        record = db.query(Video).filter(Video.vimeo_id == vimeo_id).first()

        if record and record.mux_asset_id:
            migration_status = "Migrated"
        elif record:
            migration_status = "Processing"
        else:
            migration_status = "Pending"

        rows.append({
            "Vimeo ID": vimeo_id,
            "Vimeo Title": v.get("name", "Untitled"),
            "Vimeo Folder Path": folder_name,
            "Vimeo URL": v.get("link", f"https://vimeo.com/{vimeo_id}"),
            "Mux Asset ID": record.mux_asset_id if record else "",
            "Mux Playback ID": record.mux_playback_id if record else "",
            "Mux Player URL": f"https://player.mux.com/{record.mux_playback_id}" if record and record.mux_playback_id else "",
            "Mux Stream URL": record.mux_stream_url if record else "",
            "Captions Count": record.captions_count if record else "",
            "Captions Languages": record.captions_languages if record else "",
            "Audio Tracks Count": record.audio_tracks_count if record else "",
            "Audio Languages": record.audio_languages if record else "",
            "Migrated At": record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record and record.created_at else "",
            "Migration Status": migration_status,
        })

    df_all = pd.DataFrame(rows)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_verify_sheet(writer, df_all, "All Videos")
        for folder_name in sorted(df_all["Vimeo Folder Path"].unique()):
            df_folder = df_all[df_all["Vimeo Folder Path"] == folder_name].copy()
            sheet_name = folder_name.translate(str.maketrans("", "", r'\/:*?[]'))[:31]
            _write_verify_sheet(writer, df_folder, sheet_name)

    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=verify_folder_{folder_id}.xlsx"},
    )


async def _run_drm_upgrade():
    """Background task — processes all videos, logs progress to app.log."""
    import logging
    from app.database.models import Video
    from app.services.mux_service import get_asset

    log = logging.getLogger("drm_upgrade")
    log.info("[DRM Upgrade] Background task started.")

    with SessionLocal() as db:
        videos = db.query(Video).filter(
            Video.mux_asset_id != None,
            Video.mux_signed_playback_id != None,
        ).all()
        video_snapshot = [(v.vimeo_id, v.mux_asset_id, v.mux_signed_playback_id) for v in videos]

    total = len(video_snapshot)
    log.info(f"[DRM Upgrade] Found {total} videos to process.")
    updated, skipped, failed = 0, 0, 0
    _task_status["drm_upgrade"] = {"status": "running", "total": total, "updated": 0, "skipped": 0, "failed": 0, "current": 0}

    for i, (vimeo_id, asset_id, stored_signed_id) in enumerate(video_snapshot, start=1):
        log.info(f"[DRM Upgrade] {i}/{total} — Vimeo ID: {vimeo_id} | Asset: {asset_id} | Signed: {stored_signed_id}")
        _task_status["drm_upgrade"]["current"] = i
        try:
            asset = await asyncio.to_thread(get_asset, asset_id)
        except Exception as asset_err:
            if "not_found" in str(asset_err).lower() or "Asset not found" in str(asset_err):
                log.warning(f"[DRM Upgrade] ⏭ Skipping {vimeo_id} — Mux asset no longer exists")
                skipped += 1
            else:
                failed += 1
                log.error(f"[DRM Upgrade] ❌ Failed to fetch asset {vimeo_id}: {asset_err}")
            continue

        existing_policies = {p["id"]: p.get("policy") for p in asset.get("playback_ids", [])}

        if stored_signed_id and existing_policies.get(stored_signed_id) == "drm":
            log.info(f"[DRM Upgrade] ⏭ Skipping {vimeo_id} — signed_id is already DRM")
            skipped += 1
            continue

        if any(pol == "drm" for pol in existing_policies.values()):
            log.info(f"[DRM Upgrade] ⏭ Skipping {vimeo_id} — asset already has a DRM playback ID")
            skipped += 1
            continue

        try:
            new_id, policy_type = await asyncio.to_thread(add_signed_playback_id, asset_id)
            with SessionLocal() as db:
                from app.database.models import Video
                video = db.query(Video).filter(Video.vimeo_id == vimeo_id).first()
                if video:
                    video.mux_drm_playback_id = new_id
                    db.commit()
            updated += 1
            _task_status["drm_upgrade"].update({"updated": updated, "skipped": skipped, "failed": failed})
            log.info(f"[DRM Upgrade] ✅ {vimeo_id} → {new_id} ({policy_type})")
        except Exception as e:
            failed += 1
            _task_status["drm_upgrade"].update({"updated": updated, "skipped": skipped, "failed": failed})
            log.error(f"[DRM Upgrade] ❌ Failed for {vimeo_id}: {str(e)}")

    _task_status["drm_upgrade"]["status"] = "done"
    log.info(f"[DRM Upgrade] Done. Total: {total} | Upgraded: {updated} | Skipped: {skipped} | Failed: {failed}")


@router.post("/upgrade-to-drm")
async def upgrade_playback_ids_to_drm():
    """
    Queues a background DRM upgrade for all migrated videos.
    Returns immediately — check /migration/task-status for progress.
    Safe to re-run — skips videos already on DRM playback IDs.
    """
    if not DRM_CONFIGURATION_ID:
        raise HTTPException(status_code=400, detail="DRM_CONFIGURATION_ID is not configured on this server.")

    with SessionLocal() as db:
        from app.database.models import Video
        total = db.query(Video).filter(
            Video.mux_asset_id != None,
            Video.mux_signed_playback_id != None,
        ).count()

    asyncio.create_task(_run_drm_upgrade())
    return {
        "status": "queued",
        "message": f"DRM upgrade started in background for {total} videos. Poll /migration/task-status for progress.",
        "total_queued": total,
    }


async def _run_repair_signed():
    """
    Re-adds a `signed` playback ID to every asset whose mux_signed_playback_id
    no longer exists on Mux (deleted during a failed DRM upgrade attempt).
    Safe to re-run — skips assets that already have a valid signed or DRM playback ID.
    """
    import logging
    from app.database.models import Video
    from app.services.mux_service import get_asset, add_signed_playback_id

    log = logging.getLogger("repair_signed")
    log.info("[Repair] Background repair task started.")

    with SessionLocal() as db:
        videos = db.query(Video).filter(Video.mux_asset_id != None).all()
        snapshot = [(v.vimeo_id, v.mux_asset_id, v.mux_signed_playback_id) for v in videos]

    total = len(snapshot)
    repaired, skipped, failed = 0, 0, 0
    _task_status["repair"] = {"status": "running", "total": total, "repaired": 0, "skipped": 0, "failed": 0, "current": 0}

    for i, (vimeo_id, asset_id, stored_signed_id) in enumerate(snapshot, start=1):
        log.info(f"[Repair] {i}/{total} — {vimeo_id} | asset {asset_id}")
        _task_status["repair"]["current"] = i
        try:
            try:
                asset = await asyncio.to_thread(get_asset, asset_id)
            except Exception as e:
                if "not_found" in str(e).lower():
                    log.warning(f"[Repair] ⏭ {vimeo_id} — Mux asset gone, skipping")
                    skipped += 1
                    continue
                raise

            existing_ids = {p["id"]: p.get("policy") for p in asset.get("playback_ids", [])}

            # Already has a valid signed or DRM playback ID on Mux — nothing to repair
            if stored_signed_id and stored_signed_id in existing_ids:
                log.info(f"[Repair] ⏭ {vimeo_id} — signed ID still valid on Mux, skipping")
                skipped += 1
                continue

            # Check if any DRM playback ID already exists (upgrade succeeded before)
            has_drm = any(pol == "drm" for pol in existing_ids.values())
            if has_drm:
                log.info(f"[Repair] ⏭ {vimeo_id} — already has DRM playback ID on Mux, skipping")
                skipped += 1
                continue

            # Need to re-add a signed playback ID
            log.info(f"[Repair] Adding signed playback ID for {vimeo_id}...")
            new_id, policy_type = await asyncio.to_thread(add_signed_playback_id, asset_id)

            with SessionLocal() as db:
                video = db.query(Video).filter(Video.vimeo_id == vimeo_id).first()
                if video:
                    video.mux_drm_playback_id = new_id
                    db.commit()

            repaired += 1
            _task_status["repair"].update({"repaired": repaired, "skipped": skipped, "failed": failed})
            log.info(f"[Repair] ✅ {vimeo_id} repaired → {new_id} ({policy_type})")

        except Exception as e:
            failed += 1
            _task_status["repair"].update({"repaired": repaired, "skipped": skipped, "failed": failed})
            log.error(f"[Repair] ❌ {vimeo_id}: {e}")

    _task_status["repair"]["status"] = "done"
    log.info(f"[Repair] Done. Total: {total} | Repaired: {repaired} | Skipped: {skipped} | Failed: {failed}")


@router.post("/repair-signed-playback")
async def repair_signed_playback():
    """
    Re-adds signed playback IDs to any asset that lost its signed ID during
    a failed DRM upgrade. Safe to call multiple times.
    """
    with SessionLocal() as db:
        from app.database.models import Video
        total = db.query(Video).filter(Video.mux_asset_id != None).count()

    asyncio.create_task(_run_repair_signed())
    return {
        "status": "queued",
        "message": f"Repair started for up to {total} videos. Poll /migration/task-status for progress.",
        "total_queued": total,
    }


@router.get("/task-status")
def get_task_status():
    """Returns live progress for background tasks."""
    return {
        "drm_upgrade": _task_status.get("drm_upgrade", {"status": "not_started"}),
        "repair": _task_status.get("repair", {"status": "not_started"}),
        "bulk_audio": _task_status.get("bulk_audio", {"status": "not_started"}),
    }


def _write_verify_sheet(writer, df, sheet_name):
    import pandas as pd
    df.to_excel(writer, index=False, sheet_name=sheet_name)
    ws = writer.sheets[sheet_name]

    # Colour rows by status
    from openpyxl.styles import PatternFill
    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red    = PatternFill("solid", fgColor="FFC7CE")
    status_col = df.columns.get_loc("Migration Status") + 1  # 1-based

    for row_idx, status in enumerate(df["Migration Status"], start=2):
        fill = green if status == "Migrated" else yellow if status == "Processing" else red
        for cell in ws[row_idx]:
            cell.fill = fill

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)