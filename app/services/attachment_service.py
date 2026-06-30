import os
import re
import time
import logging
import requests
import openpyxl

from app.config import MUX_TOKEN_ID, MUX_TOKEN_SECRET, SERVER_BASE_URL

logger = logging.getLogger(__name__)

BASE_DIR       = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEMP_AUDIO_DIR = os.path.join(BASE_DIR, "temp_audio")

MUX_BASE = "https://api.mux.com/video/v1"
MUX_AUTH = (MUX_TOKEN_ID, MUX_TOKEN_SECRET)

CLEANUP_DELAY_SECONDS = 20


def _normalize(name: str) -> str:
    name = os.path.splitext(name)[0]
    name = name.lower()
    name = re.sub(r"[\s\-]+", "_", name)
    name = re.sub(r"[^a-z0-9_]", "", name)
    return name


def _find_match(video_name: str, file_map: dict[str, str]) -> str | None:
    """Return the file path whose normalized stem best matches video_name."""
    key = _normalize(video_name)
    # Exact normalized match
    if key in file_map:
        return file_map[key]
    # Prefix match (video_name is prefix of filename or vice-versa)
    for norm, path in file_map.items():
        if norm.startswith(key) or key.startswith(norm):
            return path
    return None


def read_excel_rows(excel_path: str, asset_id_col: str, name_col: str) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        asset_id  = record.get(asset_id_col)
        video_name = record.get(name_col)
        if asset_id and video_name:
            rows.append({"asset_id": str(asset_id).strip(), "video_name": str(video_name).strip()})
    wb.close()
    return rows


def _attach_track(asset_id: str, public_url: str, track_type: str, language: str, name: str):
    payload = {"url": public_url, "type": track_type, "language_code": language, "name": name}
    if track_type == "text":
        payload["text_type"] = "subtitles"
    r = requests.post(
        f"{MUX_BASE}/assets/{asset_id}/tracks",
        json=payload,
        auth=MUX_AUTH,
    )
    if not r.ok:
        raise Exception(f"Mux API error ({r.status_code}): {r.text}")
    return r.json()["data"]


def attach_tracks_to_asset(
    asset_id: str,
    audio_path: str | None,
    srt_path: str | None,
    audio_language: str,
    audio_name: str,
    srt_language: str,
) -> dict:
    result = {"audio": None, "srt": None, "errors": []}
    os.makedirs(TEMP_AUDIO_DIR, exist_ok=True)
    temp_files: list[str] = []

    try:
        if audio_path:
            filename = os.path.basename(audio_path)
            dest = os.path.join(TEMP_AUDIO_DIR, filename)
            if audio_path != dest:
                import shutil
                shutil.copy2(audio_path, dest)
            temp_files.append(dest)
            url = f"{SERVER_BASE_URL}/temp-audio/{filename}"
            try:
                _attach_track(asset_id, url, "audio", audio_language, audio_name)
                result["audio"] = "attached"
                logger.info(f"[Attachment] ✅ Audio attached to {asset_id}")
            except Exception as e:
                result["errors"].append(f"audio: {e}")
                logger.error(f"[Attachment] ❌ Audio failed for {asset_id}: {e}")

        if srt_path:
            filename = os.path.basename(srt_path)
            dest = os.path.join(TEMP_AUDIO_DIR, filename)
            if srt_path != dest:
                import shutil
                shutil.copy2(srt_path, dest)
            temp_files.append(dest)
            url = f"{SERVER_BASE_URL}/temp-audio/{filename}"
            try:
                _attach_track(asset_id, url, "text", srt_language, srt_language)
                result["srt"] = "attached"
                logger.info(f"[Attachment] ✅ SRT attached to {asset_id}")
            except Exception as e:
                result["errors"].append(f"srt: {e}")
                logger.error(f"[Attachment] ❌ SRT failed for {asset_id}: {e}")

        # Let Mux fetch the files before we delete them
        if temp_files:
            time.sleep(CLEANUP_DELAY_SECONDS)

    finally:
        for f in temp_files:
            try:
                if os.path.exists(f):
                    os.remove(f)
            except Exception:
                pass

    return result
